import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "一周蔬菜方案(600g)"

header_font = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
day_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
day_fills = {
    "周一": PatternFill(start_color="7B2D8B", end_color="7B2D8B", fill_type="solid"),
    "周二": PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid"),
    "周三": PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "周四": PatternFill(start_color="C62828", end_color="C62828", fill_type="solid"),
    "周五": PatternFill(start_color="6D4C41", end_color="6D4C41", fill_type="solid"),
    "周六": PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid"),
    "周日": PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid"),
}
cell_font = Font(name="Microsoft YaHei", size=10)
total_font = Font(name="Microsoft YaHei", bold=True, size=10, color="C62828")
total_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["星期", "主题", "食材", "克重(g)", "烹饪方式", "功效亮点", "绿叶菜标记"]
col_widths = [8, 16, 14, 10, 22, 30, 12]

for col_i, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=1, column=col_i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border
    ws.column_dimensions[get_column_letter(col_i)].width = w

plan = [
    ("周一", "紫色抗氧化日", [
        ("紫甘蓝",   70,  "生切丝凉拌",            "花青素、维C",                          ""),
        ("油麦菜",   90,  "焯水30秒",              "维C、叶酸丰富，非十字花科",             "✔"),
        ("西兰花",  170,  "蒸5分钟",               "萝卜硫素（十字花科须熟食）",             ""),
        ("番茄",    160,  "煮汤",                  "番茄红素，煮后吸收更好",                ""),
        ("白蘑菇",  110,  "空气炸锅 180°C/8min",   "硒、维D，利于甲状腺",                   ""),
    ]),
    ("周二", "深绿修复日", [
        ("菠菜",    180,  "焯水30秒",              "叶酸、镁、抗炎",                        "✔"),
        ("彩椒(红+黄)", 160, "生吃切条",            "维C含量极高",                           ""),
        ("胡萝卜",  140,  "蒸10分钟",              "β-胡萝卜素，蒸后吸收↑",                 ""),
        ("秋葵",    120,  "水煮2分钟",             "黏液多糖、抗炎",                        ""),
    ]),
    ("周三", "橙黄免疫日", [
        ("南瓜",    210,  "蒸15分钟",              "β-胡萝卜素、维A",                       ""),
        ("芦笋",    130,  "空气炸锅 190°C/6min",   "谷胱甘肽、抗氧化",                      ""),
        ("黄瓜",     70,  "生吃",                  "补水、低热量",                          ""),
        ("茼蒿",    100,  "焯水1分钟",             "类黄酮、维K，抗炎突出",                  "✔"),
        ("香菇",     90,  "煮汤",                  "硒、β-葡聚糖、调节免疫",                ""),
    ]),
    ("周四", "红色番茄日", [
        ("番茄",    220,  "煮成浓汤",              "番茄红素（加热释放更多）",                ""),
        ("花椰菜",  170,  "蒸6分钟",               "抗氧化（十字花科须熟食）",               ""),
        ("紫洋葱",  100,  "空气炸锅 180°C/10min",  "槲皮素、抗炎",                          ""),
        ("生菜",    110,  "手撕生吃",              "膳食纤维、维K",                         "✔"),
    ]),
    ("周五", "菌菇修护日", [
        ("杏鲍菇",  160,  "空气炸锅 190°C/8min",   "口感似肉、富硒",                        ""),
        ("西兰花",  170,  "水煮3分钟",             "萝卜硫素（十字花科须熟食）",             ""),
        ("小番茄",  120,  "生吃",                  "维C、番茄红素",                         ""),
        ("空心菜",  150,  "焯水30秒",              "铁、叶绿素含量高，抗氧化",              "✔"),
    ]),
    ("周六", "绿紫双拼日", [
        ("紫薯",    170,  "蒸20分钟",              "花青素极高",                            ""),
        ("小白菜",  150,  "蒸3分钟",               "钙、维K出色（十字花科须熟食）",          "✔"),
        ("甜椒(红)", 120, "生吃",                  "维C是橙子的3倍",                        ""),
        ("茄子",    160,  "空气炸锅 200°C/10min",  "茄色素、无需油炸也酥脆",                ""),
    ]),
    ("周日", "彩虹综合日", [
        ("羽衣甘蓝", 130, "焯水1分钟（不要生吃）", "维K、叶黄素（十字花科须熟食）",          "✔"),
        ("胡萝卜",  130,  "空气炸条 190°C/8min",   "β-胡萝卜素",                            ""),
        ("番茄",    190,  "蒸",                    "番茄红素",                              ""),
        ("彩椒(黄)", 150, "生切片",                "维C、类黄酮",                           ""),
    ]),
]

row = 2
for day, theme, items in plan:
    start_row = row
    day_total = 0
    for name, grams, cook, benefit, leaf in items:
        ws.cell(row=row, column=1, value=day).font = day_font
        ws.cell(row=row, column=1).fill = day_fills[day]
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=theme).font = cell_font
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=name).font = cell_font
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=grams).font = cell_font
        ws.cell(row=row, column=4).alignment = center
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=cook).font = cell_font
        ws.cell(row=row, column=5).alignment = left_wrap
        ws.cell(row=row, column=5).border = thin_border

        ws.cell(row=row, column=6, value=benefit).font = cell_font
        ws.cell(row=row, column=6).alignment = left_wrap
        ws.cell(row=row, column=6).border = thin_border

        ws.cell(row=row, column=7, value=leaf).font = cell_font
        ws.cell(row=row, column=7).alignment = center
        ws.cell(row=row, column=7).border = thin_border
        if leaf == "✔":
            ws.cell(row=row, column=7).font = Font(name="Microsoft YaHei", size=10, color="2E7D32", bold=True)

        day_total += grams
        row += 1

    end_row = row - 1
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)

    tc = ws.cell(row=row, column=3, value="当日合计")
    tc.font = total_font
    tc.fill = total_fill
    tc.alignment = center
    tc.border = thin_border

    tv = ws.cell(row=row, column=4, value=day_total)
    tv.font = total_font
    tv.fill = total_fill
    tv.alignment = center
    tv.border = thin_border

    for ci in [1, 2, 5, 6, 7]:
        sc = ws.cell(row=row, column=ci)
        sc.fill = total_fill
        sc.border = thin_border
    row += 1

row += 1
ws.cell(row=row, column=1, value="桥本专项备注").font = Font(name="Microsoft YaHei", bold=True, size=11, color="4472C4")
row += 1
notes = [
    ("十字花科必须熟食", "西兰花、花椰菜、紫甘蓝、羽衣甘蓝、小白菜——蒸/煮后致甲状腺肿素降解85%+，安全食用"),
    ("硒很重要", "每周安排3-4次菌菇（香菇、白蘑菇、杏鲍菇），硒支持甲状腺过氧化物酶"),
    ("避免大量生食十字花科", "方案中所有十字花科均标注为蒸或水煮，紫甘蓝少量凉拌可接受（70g以内）"),
    ("碘摄入适量即可", "海带紫菜等高碘食物本方案未纳入，桥本患者碘宜适中，不必刻意补"),
    ("搭配优质脂肪", "蒸/煮蔬菜淋少许初榨橄榄油或亚麻籽油，帮助脂溶性抗氧化物吸收"),
    ("每日绿叶菜保证", "调整后每天至少含一种绿叶菜，一周7种不重复，周均绿叶菜约980g"),
]
for key, val in notes:
    ws.cell(row=row, column=1, value=key).font = Font(name="Microsoft YaHei", bold=True, size=10)
    ws.cell(row=row, column=1).alignment = left_wrap
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    ws.cell(row=row, column=2, value=val).font = cell_font
    ws.cell(row=row, column=2).alignment = left_wrap
    row += 1

summary_ws = wb.create_sheet("绿叶菜周总览")
s_headers = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
leafy = [
    ("油麦菜", "90g"),
    ("菠菜", "180g"),
    ("茼蒿", "100g"),
    ("生菜", "110g"),
    ("空心菜", "150g"),
    ("小白菜", "150g"),
    ("羽衣甘蓝", "130g"),
]
for ci, h in enumerate(s_headers, 1):
    c = summary_ws.cell(row=1, column=ci, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border
    summary_ws.column_dimensions[get_column_letter(ci)].width = 16

for ci, (name, g) in enumerate(leafy, 1):
    c = summary_ws.cell(row=2, column=ci, value=f"{name} {g}")
    c.font = Font(name="Microsoft YaHei", size=11, color="2E7D32", bold=True)
    c.alignment = center
    c.border = thin_border

summary_ws.cell(row=4, column=1, value="周绿叶菜总量").font = Font(name="Microsoft YaHei", bold=True, size=11)
summary_ws.cell(row=4, column=2, value="910g").font = Font(name="Microsoft YaHei", bold=True, size=11, color="C62828")

outpath = "/opt/feishu-bridge/桥本友好_一周蔬菜方案_600g.xlsx"
wb.save(outpath)
print(f"Excel saved: {outpath}")
